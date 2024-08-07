import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from fuzzywuzzy import fuzz, process
import io
import streamlit as st


column_mappings = {
    'Isin Code': [
        'ISIN', 'Isin', 'Share ISIN Reference', 'FINANZINSTRUMENT_IDENT', 'Text23'
    ],
    'Provision': [
        'Comm. Amount', 'Provision', 'Betrag (€)', 'Vergütung', 'EURMonat',
        'Client Trailer Fees Amount In Consolidated Currency', 'Amount In Agreement Ccy', 'Bepro',
        'Provisionsbetrag in Währung', 'Fee', 'BPROV', 'BpkEUR', 'Kommissionsbetrag', 'Commission Due Payment CCY', 'Fee (Payment Currency)', 'Amount In Partner Currency',
        'Betrag (EUR)'
    ],
    'Date': [
        'Date', 'Datum', 'Booking Date', 'End-Datum', 'Period End Date', 'Holding as of', 'STICHTAG', 'Period', 'Datum_str', 'Positionsdatum', 'Stichtag', 'Retrocession Date'
    ]
}

def read_files_from_upload(uploaded_files):
    """Read all uploaded files and return a dictionary of DataFrames."""
    files_data = {}
    for uploaded_file in uploaded_files:
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            all_sheets_df = pd.DataFrame()
            header_set = False
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
                df = date_converter(df)
                if not header_set:
                    df = set_correct_headers(df, column_mappings)
                    header_set = True
                else:
                    df.columns = all_sheets_df.columns
                all_sheets_df = pd.concat([all_sheets_df, df], ignore_index=True)
                logging.info(f"Read {uploaded_file.name} - {sheet_name} successfully with shape {df.shape}")
            files_data[uploaded_file.name] = all_sheets_df
            logging.info(f"Combined all sheets for {uploaded_file.name} with shape {all_sheets_df.shape}")
        except Exception as e:
            logging.error(f"Error reading {uploaded_file.name}: {e}")
    return files_data

def date_converter(df):
    if "Statement Month" in df.columns and "Statement Year" in df.columns:
        df["Date"] = pd.to_datetime(df["Statement Month"].astype(str) + '-' + df["Statement Year"].astype(str), format='%m-%Y')
    return df

def set_correct_headers(df, column_mappings):
    """Set the correct headers by finding the row that contains any of the mapped column names."""
    for i, row in df.iterrows():
        if any(header in row.values for headers in column_mappings.values() for header in headers):
            df.columns = df.iloc[i]
            df = df.drop(i).reset_index(drop=True)
            return df
    return df

def rename_columns(df, column_mappings):
    """Rename columns in the dataframe based on the provided mappings."""
    rename_dict = {}
    for new_name, old_names in column_mappings.items():
        for old_name in old_names:
            if old_name in df.columns:
                rename_dict[old_name] = new_name
    df.rename(columns=rename_dict, inplace=True)
    return df

def filter_valid_rows(df):
    """Filter rows that have both a valid ISIN code and a valid Date."""
    if 'Isin Code' in df.columns and 'Date' in df.columns:
        df = df.dropna(subset=['Isin Code', 'Date'])
        df = df[df['Isin Code'].apply(lambda x: isinstance(x, str) and x.strip() != "")]
    return df

def convert_date_column(df, date_column_name):
    if date_column_name in df.columns:
        df[date_column_name] = pd.to_datetime(df[date_column_name], errors='coerce', format='%d.%m.%Y')
    return df

def aggregate_data(data, column_name):
    """Aggregate data by ISIN and sum the specified column."""
    aggregated_data = {}
    for file_name, df in data.items():
        df = rename_columns(df, column_mappings)
        df = filter_valid_rows(df)
        if 'Isin Code' in df.columns and column_name in df.columns:
            aggregated_df = df.groupby('Isin Code')[column_name].sum().reset_index()
            aggregated_data[file_name] = aggregated_df
            #logging.info(f"Aggregated {file_name} successfully with shape {aggregated_df.shape}")
        else:
            #logging.info(f"Required columns not found in {file_name}")
    return aggregated_data

def match_files(fundline_files, excel_files):
    """Match Fundline and Excel files based on similar names using fuzzy matching."""
    matched_files = []
    for fundline_file in fundline_files:
        fundline_base = os.path.splitext(fundline_file)[0].lower()
        best_match, score = process.extractOne(fundline_base, [os.path.splitext(f)[0].lower() for f in excel_files], scorer=fuzz.partial_ratio)
        if score > 80:  # Adjust this threshold based on your needs
            excel_file = next(f for f in excel_files if os.path.splitext(f)[0].lower() == best_match)
            matched_files.append((fundline_file, excel_file))
    logging.info("Matched files: %s", matched_files)
    return matched_files

def compare_data(fundline_data, excel_data, column_mappings):
    """Compare data from matched Fundline and Excel files based on ISIN and Date and create a DataFrame showing differences."""
    comparison_files = []

    fundline_files = fundline_data.keys()
    excel_files = excel_data.keys()
    matched_files = match_files(fundline_files, excel_files)

    for fundline_file, excel_file in matched_files:
        fundline_df = fundline_data[fundline_file]
        excel_df = excel_data[excel_file]

        # Rename columns
        fundline_df = rename_columns(fundline_df, column_mappings)
        excel_df = rename_columns(excel_df, column_mappings)

        # Convert date columns to datetime64[ns]
        fundline_df = convert_date_column(fundline_df, 'Date')
        excel_df = convert_date_column(excel_df, 'Date')

        fundline_df = fundline_df.groupby(['Isin Code', 'Date'])['Erwartete Prov. Whg'].sum().reset_index()
        excel_df = excel_df.groupby(['Isin Code', 'Date'])['Provision'].sum().reset_index()

        # Check if required columns are present
        if 'Isin Code' in fundline_df.columns and 'Date' in fundline_df.columns and 'Isin Code' in excel_df.columns and 'Date' in excel_df.columns:
            # Merge dataframes on ISIN and Date with suffixes to handle common columns
            comparison_df = pd.merge(
                fundline_df, excel_df, 
                left_on=['Isin Code', 'Date'], 
                right_on=['Isin Code', 'Date'], 
                how='inner', 
                suffixes=('_Fundline', '_Excel')
            )

            # Identify the correct column names
            fundline_column = 'Erwartete Prov. Whg_Fundline' if 'Erwartete Prov. Whg_Fundline' in comparison_df.columns else 'Erwartete Prov. Whg'
            excel_column = 'Provision_Excel' if 'Provision_Excel' in comparison_df.columns else 'Provision'

            # Ensure the columns being compared have the same type
            comparison_df[fundline_column] = comparison_df[fundline_column].astype(float)
            comparison_df[excel_column] = comparison_df[excel_column].astype(float)

            # Calculate differences and add a new column for the difference
            comparison_df['Difference'] = comparison_df[excel_column] - comparison_df[fundline_column]

            # Aggregate results for Quartal
            fundline_quartal_agg = fundline_df.groupby('Isin Code')['Erwartete Prov. Whg'].sum().reset_index()
            excel_quartal_agg = excel_df.groupby('Isin Code')['Provision'].sum().reset_index()
            quartal_aggregated_df = pd.merge(
                fundline_quartal_agg, excel_quartal_agg, 
                on='Isin Code', 
                how='inner', 
                suffixes=('_Fundline', '_Excel')
            )
            quartal_aggregated_df['Difference'] = quartal_aggregated_df['Provision'] - quartal_aggregated_df['Erwartete Prov. Whg']

            # Save each comparison result to a file
            if not comparison_df.empty:
                output = io.BytesIO()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    quartal_aggregated_df.to_excel(writer, sheet_name='Quartal', index=False)
                    comparison_df[['Isin Code', 'Date', fundline_column, excel_column, 'Difference']].to_excel(writer, sheet_name='Einzeln', index=False)

                #appyling formatting if they are below or up certain value
                apply_conditional_formatting(output, sheet_name='Quartal', column='D', lower_threshold=-20, upper_threshold=20)
                apply_conditional_formatting(output, sheet_name='Einzeln', column='E', lower_threshold=-20, upper_threshold=20)

                output.seek(0)
                comparison_files.append((f"{os.path.splitext(fundline_file)[0]}_{os.path.splitext(excel_file)[0]}_comparison.xlsx", output))
                logging.info(f"Saved comparison results to in-memory file for {fundline_file} and {excel_file}")
        else:
            logging.info(f"Required columns not found in {fundline_file} or {excel_file}")

    return comparison_files

def apply_conditional_formatting(output, sheet_name='Sheet2', column='Difference', lower_threshold=None, upper_threshold=None):
    """Apply conditional formatting to the Excel file for the specified column and range."""
    output.seek(0)
    wb = load_workbook(output)
    ws = wb[sheet_name]

    end_row = ws.max_row

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    if lower_threshold is not None:
        lower_rule = CellIsRule(operator='lessThan', formula=[str(lower_threshold)], fill=red_fill)
        ws.conditional_formatting.add(f'{column}2:{column}{end_row}', lower_rule)

    if upper_threshold is not None:
        upper_rule = CellIsRule(operator='greaterThan', formula=[str(upper_threshold)], fill=red_fill)
        ws.conditional_formatting.add(f'{column}2:{column}{end_row}', upper_rule)

    output.seek(0)
    wb.save(output)

# Streamlit UI
st.set_page_config(
    page_title="Data Comparison Tool for RSC",
    layout="centered",
    page_icon="https://upload.wikimedia.org/wikipedia/de/e/eb/Raiffeisen_%C3%96sterreich_logo.svg",
)

st.header("Data Comparison Tool")

st.sidebar.header("Upload Files")

fundline_files = st.sidebar.file_uploader(
    "Upload Fundline files", type=["xlsx"], accept_multiple_files=True
)
excel_files = st.sidebar.file_uploader(
    "Upload Excel files", type=["xlsx"], accept_multiple_files=True
)

if st.sidebar.button("Run Comparison"):
    if not fundline_files or not excel_files:
        st.error("Please upload files before running the comparison.")
    else:
        # Read uploaded files
        fundline_data = read_files_from_upload(fundline_files)
        excel_data = read_files_from_upload(excel_files)

        # Perform comparisons
        comparison_files = compare_data(fundline_data, excel_data, column_mappings)

        # Show results
        if comparison_files:
            st.write("Comparison results:")
            for file_name, file_data in comparison_files:
                st.download_button(
                    label=f"Download {file_name}", data=file_data, file_name=file_name
                )
        else:
            st.write("No discrepancies found.")

